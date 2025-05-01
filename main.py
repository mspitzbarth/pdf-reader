import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
import pdfplumber
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule
import warnings
import threading
import subprocess

warnings.filterwarnings("ignore")


# ========= PDF Tabellenextraktor =========
def extract_tables_from_pdf(pdf_path):
    all_tables = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            tables = page.extract_tables()
            for table in tables:
                cleaned = [row for row in table if row and len(row) > 3 and "Report erstellt" not in row[0]]
                if cleaned:
                    header = cleaned[0]
                    rows = cleaned[1:]
                    df = pd.DataFrame(rows, columns=header)

                    expected_cols = [
                        "Datum", 
                        "Archiv.Biogasanlage.VL PRI", 
                        "Archiv.Biogasanlage.Pumpe Störung", 
                        "Archiv.Wärmezähler.Biogasanlage.rZaehlerstand_kommuliert"
                    ]
                    if len(df.columns) >= 4:
                        df = df.iloc[:, :4]
                        df.columns = expected_cols
                        df["Datum"] = pd.to_datetime(df["Datum"], errors="coerce")
                        df["Archiv.Biogasanlage.VL PRI"] = pd.to_numeric(df["Archiv.Biogasanlage.VL PRI"], errors="coerce")
                        df["Archiv.Biogasanlage.Pumpe Störung"] = df["Archiv.Biogasanlage.Pumpe Störung"].astype(str).str.lower().map({"true": True, "false": False})
                        df["Archiv.Wärmezähler.Biogasanlage.rZaehlerstand_kommuliert"] = pd.to_numeric(df["Archiv.Wärmezähler.Biogasanlage.rZaehlerstand_kommuliert"], errors="coerce")
                        all_tables.append(df)

    return pd.concat(all_tables, ignore_index=True) if all_tables else None

# ========= Bedingte Formatierung =========
def apply_conditional_formatting(excel_path, sheet_name="Daten", value_col_letter="B", greater_than_value=80, less_than_value=80):
    wb = load_workbook(excel_path)
    ws = wb[sheet_name]

    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    value_range = f"{value_col_letter}2:{value_col_letter}{ws.max_row}"
    ws.conditional_formatting.add(value_range, CellIsRule(operator='greaterThan', formula=[str(greater_than_value)], fill=green_fill))
    ws.conditional_formatting.add(value_range, CellIsRule(operator='lessThanOrEqual', formula=[str(less_than_value)], fill=red_fill))

    wb.save(excel_path)

# ========= Speichern =========
def save_with_stats(df, output_path, stats, greater_than_value, less_than_value):
    total = stats["Gesamte Zeilen"]
    above_count = stats[f"Werte > {greater_than_value}"]
    below_count = stats[f"Werte ≤ {greater_than_value}"]

    # Now include percentage calculation
    enhanced_stats = {
        "Gesamte Zeilen (Anzahl)": total,
        f"Werte > {greater_than_value} (Anzahl)": above_count,
        f"Werte > {greater_than_value} (%)": f"{(above_count / total) * 100:.2f}%",
        f"Werte ≤ {greater_than_value} (Anzahl)": below_count,
        f"Werte ≤ {greater_than_value} (%)": f"{(below_count / total) * 100:.2f}%",
    }

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name="Daten", index=False)
        stats_df = pd.DataFrame(list(enhanced_stats.items()), columns=["Bezeichnung", "Wert"])
        stats_df.to_excel(writer, sheet_name="Gesamtergebnis", index=False)

    apply_conditional_formatting(output_path, greater_than_value=greater_than_value, less_than_value=less_than_value)

# ========= Hauptanwendung =========
class PDFExtractorApp(tk.Tk):
    def __init__(self):
        super().__init__()

        self.title("PDF Tabellen Extraktor")
        self.geometry("1100x800")
        self.configure(bg="#f0f0f0")

        # ---- Treeview Styling ----
        style = ttk.Style(self)
        style.theme_use("clam")
        style.configure("Treeview",
                        background="#ffffff",
                        foreground="#000000",
                        rowheight=30,
                        fieldbackground="#ffffff",
                        font=("Arial", 10))
        style.configure("Treeview.Heading",
                        background="#d9d9d9",
                        foreground="#000000",
                        font=("Arial", 11, "bold"))
        style.configure("TNotebook", background="#f0f0f0", borderwidth=0)
        style.configure("TNotebook.Tab", font=("Arial", 11), padding=[10, 5])

        self.resizable(False, False)

        self.pdf_path = None
        self.xlsx_path = None
        self.df = None

        self.create_widgets()

    def create_widgets(self):
        # == Top File Selection ==
        file_frame = tk.Frame(self, bg="#f0f0f0")
        file_frame.pack(pady=20)

        self.select_button = tk.Button(file_frame, text="📄 PDF auswählen", command=self.select_file,
                                    width=25, font=("Arial", 12), bg="white", fg="black", relief="solid", bd=1)
        self.select_button.grid(row=0, column=0, padx=10, pady=5)

        tk.Label(file_frame, text="Schwellwert:", bg="#f0f0f0", fg="black", font=("Arial", 12, "bold")).grid(row=0, column=1, padx=(30, 5), pady=5)

        self.threshold_entry = tk.Entry(file_frame, font=("Arial", 12), width=8, justify="center", bg="white", relief="solid", bd=1, fg="black")
        self.threshold_entry.grid(row=0, column=2, padx=5, pady=5)
        self.threshold_entry.insert(0, "80")

        # == Action Buttons ==
        action_frame = tk.Frame(self, bg="#f0f0f0")
        action_frame.pack(pady=20)

        self.start_button = tk.Button(action_frame, text="▶ Start", command=self.start_processing, state="disabled",
                                    width=20, font=("Arial", 12), bg="white", fg="black", relief="solid", bd=1)
        self.start_button.grid(row=0, column=0, padx=20, pady=5)

        self.open_button = tk.Button(action_frame, text="📊 Excel öffnen", command=self.open_file, state="disabled",
                                    width=20, font=("Arial", 12), bg="white", fg="black", relief="solid", bd=1)
        self.open_button.grid(row=0, column=1, padx=20, pady=5)

        # == Progress Bar ==
        progress_frame = tk.Frame(self, bg="#f0f0f0")
        progress_frame.pack(pady=10)

        self.progress = ttk.Progressbar(progress_frame, orient="horizontal", mode="determinate", length=600)
        self.progress.pack(pady=5)

        self.progress_label = tk.Label(progress_frame, text="Noch keine Datei geladen",
                                    bg="#f0f0f0", fg="black", font=("Arial", 10))
        self.progress_label.pack()

        # == Notebook Tabs ==
        self.notebook = ttk.Notebook(self)
        self.notebook.pack(expand=True, fill="both", pady=10, padx=20)

        self.tabs = {}
        for tab_name in ["Gesamte Zeilen", "Werte > Schwellwert", "Werte ≤ Schwellwert"]:
            frame = tk.Frame(self.notebook, bg="#ffffff")
            self.notebook.add(frame, text=tab_name)
            self.tabs[tab_name] = frame

        # == Treeviews in each tab ==
        self.trees = {}
        self.count_labels = {}
        for key, frame in self.tabs.items():
            columns = [
                "Datum",
                "Archiv.Biogasanlage.VL PRI",
                "Archiv.Biogasanlage.Pumpe Störung",
                "Archiv.Wärmezähler.Biogasanlage.rZaehlerstand_kommuliert"
            ]
            tree = ttk.Treeview(frame, columns=columns, show="headings")
            for col in columns:
                tree.heading(col, text=col)
                tree.column(col, width=220, anchor="center")
            tree.pack(side="top", fill="both", expand=True, pady=(5, 0), padx=5)

            scrollbar = ttk.Scrollbar(frame, orient="vertical", command=tree.yview)
            tree.configure(yscrollcommand=scrollbar.set)
            scrollbar.pack(side="right", fill="y")

            count_label = tk.Label(frame, text="", font=("Arial", 10), bg="#ffffff", fg="black")
            count_label.pack(side="bottom", pady=8)

            # Add zebra striping tags
            tree.tag_configure('oddrow', background='#f9f9f9')
            tree.tag_configure('evenrow', background='#ffffff')

            self.trees[key] = tree
            self.count_labels[key] = count_label

    # Datei auswählen
    def select_file(self):
        path = filedialog.askopenfilename(title="PDF Datei auswählen", filetypes=[("PDF Dateien", "*.pdf")])
        if path:
            self.pdf_path = path
            self.start_button.config(state="normal")
            self.progress_label.config(text="Datei ausgewählt. Bereit zum Start.")

    # Verarbeitung starten
    def start_processing(self):
        if not self.pdf_path:
            return

        try:
            threshold = float(self.threshold_entry.get())
        except ValueError:
            messagebox.showerror("Ungültige Eingabe", "Schwellwert muss eine Zahl sein.")
            return

        self.start_button.config(state="disabled")
        self.open_button.config(state="disabled")
        self.progress_label.config(text="Verarbeitung läuft...")

        self.progress["mode"] = "determinate"
        self.progress["value"] = 0

        def worker():
            self.df = extract_tables_from_pdf(self.pdf_path)

            if self.df is not None:
                total_rows = len(self.df)
                above = self.df[self.df["Archiv.Biogasanlage.VL PRI"] > threshold]
                below = self.df[self.df["Archiv.Biogasanlage.VL PRI"] <= threshold]

                for tree in self.trees.values():
                    tree.delete(*tree.get_children())

                for idx, row in enumerate(self.df.iterrows(), start=1):
                    _, data = row
                    tag = 'evenrow' if idx % 2 == 0 else 'oddrow'
                    self.trees["Gesamte Zeilen"].insert("", "end", values=list(data), tags=(tag,))

                for i, (_, data) in enumerate(above.iterrows(), start=1):
                    tag = 'evenrow' if i % 2 == 0 else 'oddrow'
                    self.trees["Werte > Schwellwert"].insert("", "end", values=list(data), tags=(tag,))

                for i, (_, data) in enumerate(below.iterrows(), start=1):
                    tag = 'evenrow' if i % 2 == 0 else 'oddrow'
                    self.trees["Werte ≤ Schwellwert"].insert("", "end", values=list(data), tags=(tag,))

                stats = {
                    "Gesamte Zeilen": total_rows,
                    f"Werte > {threshold}": len(above),
                    f"Werte ≤ {threshold}": len(below)
                }

                self.count_labels["Gesamte Zeilen"].config(text=f"Zeilen insgesamt: {total_rows}")
                self.count_labels["Werte > Schwellwert"].config(text=f"Werte > {threshold}: {len(above)}")
                self.count_labels["Werte ≤ Schwellwert"].config(text=f"Werte ≤ {threshold}: {len(below)}")

                self.progress["value"] = 100

                # Save and apply conditional formatting
                output_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Dateien", "*.xlsx")])
                if output_path:
                    save_with_stats(self.df, output_path, stats, threshold, threshold)

                self.progress_label.config(text="Verarbeitung abgeschlossen.")
                self.start_button.config(state="normal")
                self.open_button.config(state="normal")
        threading.Thread(target=worker).start()

    # Excel Datei öffnen
    def open_file(self):
        if self.df is not None:
            output_path = filedialog.askopenfilename(defaultextension=".xlsx", filetypes=[("Excel Dateien", "*.xlsx")])
            if output_path:
                subprocess.Popen(["start", output_path], shell=True)

# Anwendung starten
if __name__ == "__main__":
    app = PDFExtractorApp()
    app.mainloop()
